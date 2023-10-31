import os
import pathlib
import sys
import tempfile
from datetime import datetime, timedelta
from typing import List

import openpyxl
import pandas as pd
from openpyxl.drawing.image import Image


def get_current_week(today: datetime | None = None):
    today = today if today else datetime.now()

    match today.weekday():
        case 6:
            delta = 0
        case _:
            delta = today.weekday() + 1

    sunday = today - timedelta(days=delta)
    return [sunday + timedelta(days=i) for i in range(0, 7)]


def get_clients() -> List[str]:
    return [pathlib.Path(client).stem for client in os.listdir("clients")]


class Parser:
    def __init__(
        self, template: pathlib.Path, client_name: str, today=datetime | None
    ) -> None:
        self.today = today
        self.workbook = openpyxl.load_workbook(template)
        self.client_name = client_name
        self.tmpdir = None
        self.filename = ""

    def fill(self):
        days = "EFGHIJK"
        sheet = self.workbook["Sheet1"]
        week = get_current_week(self.today)
        for l, d in zip(days, week):
            sheet[f"{l}{6}"].value = d.day

        sheet["G2"].value = week[0].strftime("%m/%d/%y")
        sheet["J2"].value = week[-1].strftime("%m/%d/%y")

        csv = pd.read_excel(
            pathlib.Path("clients", f"{self.client_name}.xlsx"), header=None
        )
        arr = csv.to_numpy()

        # Fill time
        _from = "E8"
        to = "K15"
        for line, row in zip(
            arr[:8, :],
            sheet.iter_rows(
                min_row=sheet[_from].row,
                min_col=sheet[_from].column,
                max_row=sheet[to].row,
                max_col=sheet[to].column,
            ),
        ):
            for val, cell in zip(line, row):
                try:
                    cell.value = val.strftime("%H:%M")
                except AttributeError:
                    pass

        # Fill mark signs
        _from = "E16"
        to = "K43"
        for line, row in zip(
            arr[8:, :],
            sheet.iter_rows(
                min_row=sheet[_from].row,
                min_col=sheet[_from].column,
                max_row=sheet[to].row,
                max_col=sheet[to].column,
            ),
        ):
            for val, cell in zip(line, row):
                if val == 1:
                    cell.value = "✔"

        # Fill total hours
        sheet["C8"] = arr[0, 7]
        sheet["C10"] = arr[1, 7]
        sheet["C12"] = arr[2, 7]
        sheet["C14"] = arr[3, 7]

        # Fill patient name
        sheet["G3"] = self.client_name.replace(" ", ", ")

        # Add logo
        img = Image("HealthCare.png")
        img.anchor = "B2"
        sheet.add_image(img)

    def save(self) -> None:
        self.tmpdir = tempfile.TemporaryDirectory()
        self.filename = pathlib.Path(
            self.tmpdir.name,
            f"{self.client_name}.xlsx",
        )
        self.workbook.save(self.filename)

        if sys.platform == "linux":
            import subprocess

            filename = pathlib.Path(
                self.tmpdir.name,
                f"{self.client_name}.pdf",
            )

            subprocess.run(
                [
                    "soffice",
                    "--headless",
                    "--convert-to",
                    "pdf",
                    self.filename,
                    "--outdir",
                    self.tmpdir.name,
                ]
            )
            self.filename = filename

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.tmpdir.cleanup()
